from google.cloud import storage
import base64
from openpyxl import load_workbook
from io import BytesIO
from google.cloud import firestore

firestore_client = firestore.Client()


storage_client = storage.Client()

def processExcel(data, context):
  batchSize=200
  """Background Cloud Function to be triggered by Cloud Storage.
  check more in https://cloud.google.com/functions/docs/calling/storage#functions-calling-storage-python
  """
  #print('Event ID: {}'.format(context.event_id))
  
  #print('Event type: {}'.format(context.event_type))
  #print('Bucket: {}'.format(event['bucket']))
  #print('File: {}'.format(data['name']))
  #print('Metageneration: {}'.format(event['metageneration']))
  #print('Created: {}'.format(event['timeCreated']))
  #print('Updated: {}'.format(event['updated']))
  #print('Check: {}'.format(storage.notification.OBJECT_FINALIZE_EVENT_TYPE))  

  #print('Created: {}'.format(data['timeCreated'])) #this here for illustration purposes
  #print('Updated: {}'.format(data['updated']))

  
  columnList=[]
  columnList.append("crdate")
  columnList.append("ctype")
  columnList.append("itembundle")
  columnList.append("studentid")
  columnList.append("studentname")
  columnList.append("phone")
  columnList.append("email")
  columnList.append("city")
  columnList.append("state")
  columnList.append("zipcode")
  columnList.append("address")
  columnList.append("schoolname")
  columnList.append("websiteurl")
  columnList.append("couriername")
  columnList.append("couriertype")
  columnList.append("awb")
  columnList.append("shipmentstatus")
  columnList.append("deliverydate")
  columnList.append("vendor")
  columnList.append("data_vendor_date")
  columnList.append("printcompletiondate")
  columnList.append("vendordispatchdate")
  columnList.append("remark")


  blob = storage_client.get_bucket(data['bucket']).get_blob(data['name']) 

  #TODO whatever else needed with blob
  # convert to string
 #json_data = blob.download_as_string()
  #print(json_data)
  #with open('tempfile.xlsx', 'wb') as f_xlsx:
  #  f_xlsx.write(base64.b64encode(blob))
  metadata=blob.metadata
  fileIndex=int(metadata["index"])
  #print(fileIndex)


 # return 0




  wb=load_workbook(filename=BytesIO(blob.download_as_bytes()))
  sheetnames=wb.sheetnames  
  ws = wb.get_sheet_by_name(sheetnames[0])

  maxRows=ws.max_row
  minRowsByIndex=fileIndex*batchSize
  if(minRowsByIndex==0):
    minRowsByIndex=2



  maxRowsByIndex=(fileIndex+1)*batchSize
  if(maxRowsByIndex>maxRows):
    maxRowsByIndex=maxRows


  print('{0} {1} {2}'.format(minRowsByIndex,maxRowsByIndex,maxRows))
  numberOfColumns=len(columnList)
  batch = firestore_client.batch()

  for rowIndex in range(minRowsByIndex, maxRowsByIndex):
    dataObject={}
    for columnIndex in range(numberOfColumns):
      dataObject[columnList[columnIndex]]=ws.cell(row=rowIndex, column=columnIndex+1).value
    doc_ref=firestore_client.collection("Certificates").document()
    batch.set(doc_ref, dataObject)


    
  batch.commit()
  wb.close()
  
    


